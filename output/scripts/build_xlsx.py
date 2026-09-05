import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule

OUTDIR = "/tmp/claude-0/-home-user-jaypretium/06541f7f-96c0-5594-af5f-280151700bf4/scratchpad"
u = pd.read_pickle(f"{OUTDIR}/universe.pkl")
prefs = pd.read_pickle(f"{OUTDIR}/prefs.pkl")
cov = pd.read_pickle(f"{OUTDIR}/cov60.pkl")
svar = pd.read_pickle(f"{OUTDIR}/sample_var.pkl")
irisk = pd.read_pickle(f"{OUTDIR}/ideas_risk.pkl")
dart = pd.read_pickle(f"{OUTDIR}/dart_verify.pkl")
ASOF = u.attrs["asof"]
P = u.attrs["params"]
E = 1e8

FONT = "Arial"
f_norm = Font(name=FONT, size=9)
f_bold = Font(name=FONT, size=9, bold=True)
f_hdr = Font(name=FONT, size=9, bold=True, color="FFFFFF")
f_blue = Font(name=FONT, size=9, color="0000FF")
f_green = Font(name=FONT, size=9, color="008000")
f_title = Font(name=FONT, size=13, bold=True)
fill_hdr = PatternFill("solid", fgColor="1F3864")
fill_yel = PatternFill("solid", fgColor="FFFF00")
fill_sp = PatternFill("solid", fgColor="FFF2CC")
fill_grey = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
PCT = "0.0%;(0.0%);-"
NUM0 = "#,##0;(#,##0);-"
NUM1 = "#,##0.0;(#,##0.0);-"
MULT = '0.0"x"'

wb = Workbook()

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = f_hdr; cell.fill = fill_hdr; cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); cell.border = border

def setw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[L(i)].width = w

def write_df(ws, df, start_row, fmts=None, fonts=None):
    fmts = fmts or {}; fonts = fonts or {}
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=j, value=col)
    style_header(ws, start_row, len(df.columns))
    for i, (_, r) in enumerate(df.iterrows(), start_row + 1):
        for j, col in enumerate(df.columns, 1):
            v = r[col]
            if isinstance(v, (np.floating, float)) and (np.isnan(v) or np.isinf(v)):
                v = None
            elif isinstance(v, np.generic):
                v = v.item()
            c = ws.cell(row=i, column=j, value=v)
            c.font = fonts.get(col, f_norm); c.border = border
            if col in fmts:
                c.number_format = fmts[col]
    return start_row + len(df)

# ======================= README =======================
ws = wb.active; ws.title = "README"
lines = [
    ("Billionfold Korea L/S — Liquidity-Screened Universe", f_title),
    (f"Data as of {ASOF} (KRX close). Source: KRX 전종목시세 via FinanceData/marcap (github, 일별 갱신). 섹터: FinanceData/stock_master(2018 스냅샷, 커버리지 ~75%).", f_norm),
    ("", f_norm),
    ("1. 목적", f_bold),
    ("BBAS Rule(2026-07-31)과 Junior Guide(2026-09-05)의 편입비·시총·유동성 제한을 KRX 전종목에 적용해 실제 운용 가능한 한국 종목 유니버스를 정의. 특수상황(Special Situation) 종목을 하이라이트.", f_norm),
    ("", f_norm),
    ("2. 적용 룰 (BBAS 원문 → 스크린 변수)", f_bold),
    ("② 종목 편입비 제한 6% (Long/Short) → Position cap = 6% × Book", f_norm),
    ("③ Event Play 10% (RM/BBAS 사전승인, 최장 40거래일) → Event Play bucket 별도", f_norm),
    ("⑥ 시총 3천억~7천억 합산 7%, 3천억 미만 편입 불가 → 시총 = 최근 1개월(20영업일) 평균 (참고: 시총 판단은 1개월 평균 고려 가능)", f_norm),
    ("Gross-Cut Period 3거래일 + 일일 주문 ADV 20% 참여 한도(컴플라이언스, 가이드 5장) → 3일 내 청산 가능 포지션 = 20% × 3일 × ADV = 60% × ADV", f_norm),
    ("ADV = min(20D, 60D) 거래대금 평균(거래정지일 제외). Stress ADV = 50% haircut (가정) — 12개월 rolling-20D ADV 저점/현재 비율 중위값 0.50으로 실증 부합 (Universe ADV stability 열).", f_norm),
    ("참고: 최근 40거래일 ±15% 일변동 2회 이상 → Short 편입비 -4% 제한 (Short_Cap 열)", f_norm),
    ("", f_norm),
    ("3. Tier 정의", f_bold),
    ("유니버스 = 1개월 평균 시총 ≥ 3,000억 & 제외종목 아님 & Event sleeve(10일) 기준 1% 이상 사이징 가능. ADV는 유니버스 필터가 아니라 사이징 변수(다이나믹): MaxPos = min(한도, 참여율 × sleeve DTL × ADV / Book)", f_norm),
    ("Sleeve별 청산 목표일수(DTL): Core/Hedge 3일(Gross-cut 대응, 한도 6%) / Alpha 5일(한도 6%) / Event Play 10일(사전승인, 한도 10%, exit date 명시)", f_norm),
    ("L1 Full size: 500억 Book·Alpha sleeve(5일) 기준 6% 가능 (ADV_min ≥ 30억)", f_norm),
    ("L2 Mid size: 3~6% (ADV_min 15~30억) / L3 Small size: 1~3% (ADV_min 5~15억)", f_norm),
    ("L4 Event-only: Alpha 기준 1% 미만이나 Event sleeve(10일) 기준 1% 이상 (ADV_min 2.5~5억)", f_norm),
    ("X Illiquid: 10일 기준으로도 1% 미만 (ADV_min < 2.5억) / X Excluded: KONEX·SPAC·관리·거래정지 / C: 시총 3,000억 미만(BBAS ⑥)", f_norm),
    ("포트 레벨 유동성 룰(제안): DTL ≤ 3일 포지션 합계 ≥ Gross의 40% → Gross-cut·Book-cut을 유동 sleeve에서 소화. Trim 트리거: DTL이 목표의 2배를 10거래일 연속 초과 시 축소", f_norm),
    ("", f_norm),
    ("4. 시트 구성", f_bold),
    ("Params: 파란색 입력값(Book size, 참여율, 청산일수, Stress haircut, 편입비 한도). 변경 시 Universe의 MaxPos/DTL 열과 Summary가 재계산됨.", f_norm),
    ("Summary: Tier·시총 구간·시장별 종목 수, 평균 지표, 밸류에이션/ROE/ROIC 평균(CIQ 갱신 후 표시)", f_norm),
    ("Universe: 유니버스 전체(A1~B). 노란 배경 = 특수상황 종목. CIQ 열은 Capital IQ Excel Plug-in 연결 시 자동 채워짐(IFERROR로 미연결 시 공란).", f_norm),
    ("Special_Situations: 이벤트 태그 또는 정량 시그널이 있는 유니버스 종목", f_norm),
    ("PreTrade_Check: 모델 포트 입력 → 종목별 한도·sleeve DTL·다이나믹 한도·Trim, 60D 공분산 기반 marginal/component VaR(Cov 시트), 포트 레벨 Net/TOP5/7% bucket/유동 sleeve/VaR 자동 점검", f_norm),
    ("DART_Verify: 이벤트 태그 종목 검증 우선순위(P1-P3), 확인할 공시 유형, 검증 결과 입력란", f_norm),
    ("Alpha_Risk: 아이디어별 expected alpha(가정 입력) vs 리스크 예산(Solo/Component VaR), 연환산 NAV 기여, Research ROI", f_norm),
    ("Cov: 유니버스 60D 일수익률 공분산 행렬(정적) + PreTrade 가중치 행. 임의 포트의 공분산 VaR 계산에 사용", f_norm),
    ("Pref_Pairs: 우선주-보통주 괴리율 및 양 leg 유동성", f_norm),
    ("Holdco_NAV: 지주사 상장자회사 지분가치 커버리지(지분율은 입력값, 검증 필요)", f_norm),
    ("Ideas: 투자 아이디어(방향·사이징·업사이드/다운사이드 가정·리스크·반대논거)", f_norm),
    ("Excluded: 제외 사유별 종목 수, Event-Play 후보(시총<3000억이지만 유동성 충분 / 대형주지만 유동성 부족)", f_norm),
    ("CIQ_Fields: CIQ 함수 mnemonics 정의 및 Bloomberg 대체식", f_norm),
    ("", f_norm),
    ("5. 한계 / 불확실성", f_bold),
    ("(1) 재무·밸류에이션(PER/PBR/EV·EBITDA/ROE/ROIC/성장률)은 본 환경에서 Capital IQ·DART 접근 불가 → CIQ 수식으로 삽입, CIQ 연결된 Excel에서 열면 자동 채움. 미연결 시 공란/n/a.", f_norm),
    ("(2) Event_Type/Event_Note는 모델 지식 기반(2026년 중반 이전) 태그 → 반드시 DART/CIQ로 검증. Quant_Flags는 가격·거래대금 데이터에서 계산된 시그널.", f_norm),
    ("(3) ADV 20% 룰의 lookback·매수/매도 적용 여부, Event Play 예외는 가이드상 '확인 필요' 항목 → Params에서 조정 가능.", f_norm),
    ("(4) 섹터 분류는 2018 스냅샷이라 2019년 이후 상장 종목은 공란.", f_norm),
]
for i, (t, f) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=t); c.font = f; c.alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 160

# ======================= Params =======================
wp = wb.create_sheet("Params")
params = [
    ("Book size 1 (억원)", 500, "BBAS 미팅 확인 필요: 초기 Book 500억 가정 (Guide 2장)"),
    ("Book size 2 (억원)", 1000, "증액 시나리오 1,000억 (Guide 2장)"),
    ("Daily participation cap (% of ADV)", P["PART"], "회사 이메일 기준 일일 주문 ADV 20% (Guide 5장). lookback 확인 필요"),
    ("Gross-cut period (days)", P["DAYS"], "BBAS 기타: Gross-Cut Period 3거래일"),
    ("Stressed ADV haircut", P["STRESS"], "가정: 스트레스 시 거래대금 50% 감소"),
    ("Single-name position limit", P["POS_LIMIT"], "BBAS ② 종목 편입비 제한 1: ±6%"),
    ("Event Play position limit", P["EVENT_LIMIT"], "BBAS ③: ±10%, RM/BBAS 사전승인, 최장 40거래일"),
    ("Mid-cap aggregate bucket", 0.07, "BBAS ⑥: 시총 3천억~7천억 합산 7%"),
    ("Top-5 aggregate limit", 0.25, "BBAS ⑤: Long/Short 각 TOP5 합산 25%"),
    ("Min market cap (억원)", P["MCAP_MIN"] / E, "BBAS ⑥: 3천억 미만 편입 불가"),
    ("Mid-cap upper bound (억원)", P["MCAP_MID"] / E, "BBAS ⑥"),
    ("Min position size (%NAV)", P["MIN_ALPHA_POS"], "PM 확인: 1%까지 사이징 가능 → 유니버스 하한"),
    ("Net exposure limit", "Min(Gross x 10%, 15%)", "BBAS ⑦: Min(Gross×10%, 15%) — 참고용 텍스트"),
    ("Gross-cut scenario", 0.30, "BBAS ③-1: 5D VaR 2회 이탈 시 Gross 30% cut (2/4/6VAR 단계는 20%). PreTrade_Check 청산계획에 사용"),
    ("Top-2 market cap code #1", "005930", "BBAS ④: 시총 1-2위 Long 8% 한도 대상"),
    ("Top-2 market cap code #2", "000660", "BBAS ④"),
    ("1D 99% VaR limit (abs)", -0.01, "BBAS ②-1: Max(매니저 VaR, -1.0%) — 신설 펀드는 -1.0% 고정"),
    ("DTL target — Alpha sleeve (days)", P["DTL"]["alpha"], "제안: alpha 종목은 5일 내 전량 청산 가능 사이즈"),
    ("DTL target — Event Play sleeve (days)", P["DTL"]["event"], "제안: Event Play는 10일 + exit date 명시 (BBAS ③ 최장 40거래일)"),
    ("Liquid sleeve minimum (share of gross, DTL ≤ Core days)", 0.40, "제안: Gross-cut(30%)+Book-cut 중첩 대비 유동 sleeve 하한"),
    ("Trim trigger (DTL / target)", 2.0, "제안: DTL이 목표의 2배 초과 시 축소 검토 (10거래일 연속)"),
    ("Reference alpha for research ROI", 0.30, "가정: 이벤트 성공 시 +30% → MaxPos × 30% = NAV 기여"),
    ("Min NAV impact worth research (%NAV)", 0.005, "가정: 0.5%p 미만이면 리서치 ROI 낮음 (메모 7장)"),
    ("VaR z (99%)", 2.326, "1D 99% 파라메트릭"),
]
wp.cell(row=1, column=1, value="Parameters (파란색 = 입력값. 변경 시 Universe/Summary 재계산)").font = f_title
for j, h in enumerate(["Parameter", "Value", "Source / Note"], 1):
    wp.cell(row=3, column=j, value=h)
style_header(wp, 3, 3)
for i, (k, v, note) in enumerate(params, 4):
    wp.cell(row=i, column=1, value=k).font = f_norm
    c = wp.cell(row=i, column=2, value=v); c.font = f_blue; c.fill = fill_yel
    if isinstance(v, float) and v < 1:
        c.number_format = PCT
    elif isinstance(v, (int, float)):
        c.number_format = NUM0
    wp.cell(row=i, column=3, value=note).font = f_norm
setw(wp, [40, 18, 90])
PB1, PB2, PPART, PDAYS, PSTRESS, PPOS = "Params!$B$4", "Params!$B$5", "Params!$B$6", "Params!$B$7", "Params!$B$8", "Params!$B$9"
PDA, PDE, PLIQ, PTRIM, PMIN = "Params!$B$21", "Params!$B$22", "Params!$B$23", "Params!$B$24", "Params!$B$15"
PALPHA, PMINNAV, PZ = "Params!$B$25", "Params!$B$26", "Params!$B$27"
ccodes = list(cov.index); NC = len(ccodes)
CFIRST, CLAST = L(3), L(3 + NC - 1)
CROWL, CROWF = 4, 4 + NC - 1
COV_MAT = f"Cov!${CFIRST}${CROWL}:${CLAST}${CROWF}"
COV_CODES = f"Cov!$A${CROWL}:$A${CROWF}"
COV_W = f"Cov!${CFIRST}$2:${CLAST}$2"

# ======================= Universe =======================
wu = wb.create_sheet("Universe")
U = u[u["In_Universe"]].copy()
tier_order = {"L1 Full size (6% @5d)": 0, "L2 Mid size (3-6% @5d)": 1, "L3 Small size (1-3% @5d)": 2, "L4 Event-only (≥1% @10d)": 3}
U["_t"] = U["Tier"].map(tier_order)
U = U.sort_values(["_t", "Mcap"], ascending=[True, False])
strong = U["Recent_Halt"] | (U["ShareJump_250D"] >= 1) | U["New_Listing"] | (U["VolSpike"] >= 2.5) | (U["Ret_1M"] >= 0.5) | (U["Ret_1M"] <= -0.3) | (U["Event_Type"] != "") | U["Is_Pref"]
U["SpecialSit"] = np.where(strong, "Y", "")
u.loc[U.index, "SpecialSit"] = U["SpecialSit"]

cols = [
    ("Code", "Code", None), ("Name", "Name", None), ("Market", "Market", None), ("CIQ Ticker", "Ticker_CIQ", None), ("BBG Ticker", "Ticker_BBG", None),
    ("Sector (KRX)", "Sector", None), ("Tier", "Tier", None), ("Mcap Bucket", "Mcap_Bucket", None), ("Special Sit", "SpecialSit", None),
    ("Event Type (K)", "Event_Type", None), ("Event Note (K, verify)", "Event_Note", None), ("Quant Flags (data)", "Quant_Flags", None),
    ("Close (KRW)", "Close", NUM0), ("Mcap (억)", "Mcap", NUM0), ("Mcap 1M avg (억)", "Mcap_1M", NUM0),
    ("ADV20 (억)", "ADV20", NUM1), ("ADV60 (억)", "ADV60", NUM1), ("ADV min (억)", "ADV_min", NUM1),
    ("Daily turnover", "Turnover_daily", "0.00%"), ("Vol 60D ann.", "Vol60_ann", PCT), ("±15% days 40D", "Vol15_40D", NUM0), ("Short cap", "Short_Cap", PCT),
    ("Ret 1M", "Ret_1M", PCT), ("Ret 3M", "Ret_3M", PCT), ("Ret 6M", "Ret_6M", PCT), ("Ret YTD", "Ret_YTD", PCT), ("Ret 12M", "Ret_12M", PCT),
    ("% from 52wH", "Pct_from_52wH", PCT), ("Vol spike 20D/120D", "VolSpike", MULT),
    ("ADV20 12M trough (억)", "ADV20_min12M", NUM1), ("ADV stability (trough/current)", "ADV_Stability", "0.00x"), ("1D 99% VaR proxy", "VaR99_1D", PCT), ("Cov flag", "Cov_flag", None),
]
hdr = [c[0] for c in cols]
formula_cols = ["MaxPos 3D-exit (억)", "MaxPos %NAV Book1", "MaxPos %NAV Book2", "DTL 6% Book1 (days)", "DTL 6% Book2 (days)", "DTL 2% Book1 (days)", "Stress MaxPos %NAV Book1", "Stress DTL 6% Book1", "Trough MaxPos %NAV Book1", "Trough MaxPos %NAV Book2", "MaxPos Alpha(5d) %NAV Book1", "MaxPos Alpha(5d) %NAV Book2", "MaxPos Event(10d) %NAV Book1", "MaxPos Event(10d) %NAV Book2", "Trough MaxPos Alpha(5d) %NAV Book1", "DTL 1% Book1 (days)", "NAV impact @ref alpha (MaxPos Alpha×α)", "Standalone VaR at MaxPos Alpha", "Research ROI flag"]
ciq_cols = [("P/E LTM", "IQ_PE_EXCL"), ("P/E NTM", "IQ_PE_EXCL_FWD_CIQ"), ("P/BV", "IQ_PBV"), ("TEV/EBITDA LTM", "IQ_TEV_EBITDA"),
            ("ROE %", "IQ_RETURN_EQUITY"), ("ROIC % (Return on Capital)", "IQ_RETURN_CAPITAL"), ("Rev growth 1Y %", "IQ_TOTAL_REV_1YR_ANN_GROWTH"),
            ("EPS growth 1Y %", "IQ_EPS_1YR_ANN_GROWTH"), ("Div yield %", "IQ_DIVIDEND_YIELD"), ("Net debt (억)", "IQ_NET_DEBT")]
all_hdr = hdr + formula_cols + [c[0] for c in ciq_cols]
HR = 3
wu.cell(row=1, column=1, value=f"Universe — {len(U)} names (as of {ASOF}; 시총≥3,000억, Event sleeve 기준 1%+ 사이징 가능). Tier = 500억 Book·Alpha sleeve(5일) 기준 최대 사이즈. 노란 배경 = Special Situation. 초록 = Params 참조 수식(다이나믹 사이징), CIQ 열 = Capital IQ plug-in 연결 시 자동").font = f_bold
for j, h in enumerate(all_hdr, 1):
    wu.cell(row=HR, column=j, value=h)
style_header(wu, HR, len(all_hdr))
colmap = {h: L(i) for i, h in enumerate(all_hdr, 1)}
cADV = colmap["ADV min (억)"]; cTick = colmap["CIQ Ticker"]
r0 = HR + 1
for i, (_, r) in enumerate(U.iterrows()):
    row = r0 + i
    for j, (h, key, fmt) in enumerate(cols, 1):
        v = r[key]
        if key in ("Mcap", "Mcap_1M", "ADV20", "ADV60", "ADV_min", "ADV20_min12M"):
            v = v / E
        if isinstance(v, (float, np.floating)) and (np.isnan(v) or np.isinf(v)):
            v = None
        elif isinstance(v, np.generic):
            v = v.item()
        c = wu.cell(row=row, column=j, value=v); c.font = f_norm
        if fmt:
            c.number_format = fmt
    adv = f"${cADV}{row}"
    fx = [
        (f"={PPART}*{PDAYS}*{adv}", NUM1),
        (f"=MIN({PPOS},{colmap['MaxPos 3D-exit (억)']}{row}/{PB1})", PCT),
        (f"=MIN({PPOS},{colmap['MaxPos 3D-exit (억)']}{row}/{PB2})", PCT),
        (f"=IF({adv}>0,{PPOS}*{PB1}/({PPART}*{adv}),\"n/a\")", NUM1),
        (f"=IF({adv}>0,{PPOS}*{PB2}/({PPART}*{adv}),\"n/a\")", NUM1),
        (f"=IF({adv}>0,Params!$B$15*{PB1}/({PPART}*{adv}),\"n/a\")", NUM1),
        (f"=MIN({PPOS},{colmap['MaxPos 3D-exit (억)']}{row}*{PSTRESS}/{PB1})", PCT),
        (f"=IF({adv}>0,{PPOS}*{PB1}/({PPART}*{adv}*{PSTRESS}),\"n/a\")", NUM1),
        (f"=IFERROR(MIN({PPOS},{PPART}*{PDAYS}*${colmap['ADV20 12M trough (억)']}{row}/{PB1}),\"n/a\")", PCT),
        (f"=IFERROR(MIN({PPOS},{PPART}*{PDAYS}*${colmap['ADV20 12M trough (억)']}{row}/{PB2}),\"n/a\")", PCT),
        (f"=MIN({PPOS},{PPART}*{PDA}*{adv}/{PB1})", PCT),
        (f"=MIN({PPOS},{PPART}*{PDA}*{adv}/{PB2})", PCT),
        (f"=MIN(Params!$B$10,{PPART}*{PDE}*{adv}/{PB1})", PCT),
        (f"=MIN(Params!$B$10,{PPART}*{PDE}*{adv}/{PB2})", PCT),
        (f"=IFERROR(MIN({PPOS},{PPART}*{PDA}*${colmap['ADV20 12M trough (억)']}{row}/{PB1}),\"n/a\")", PCT),
        (f"=IF({adv}>0,{PMIN}*{PB1}/({PPART}*{adv}),\"n/a\")", NUM1),
        (f"=${colmap['MaxPos Alpha(5d) %NAV Book1']}{row}*{PALPHA}", "0.00%"),
        (f"=IFERROR(${colmap['MaxPos Alpha(5d) %NAV Book1']}{row}*${colmap['1D 99% VaR proxy']}{row},\"\")", "0.00%"),
        (f"=IF(${colmap['NAV impact @ref alpha (MaxPos Alpha×α)']}{row}<{PMINNAV},\"Low\",\"OK\")", "General"),
    ]
    for k, (f, fmt) in enumerate(fx):
        c = wu.cell(row=row, column=len(cols) + 1 + k, value=f); c.font = f_green; c.number_format = fmt
    for k, (h, mn) in enumerate(ciq_cols):
        col = len(cols) + len(formula_cols) + 1 + k
        c = wu.cell(row=row, column=col, value=f'=IFERROR(CIQ(${cTick}{row},"{mn}"),"")'); c.font = f_norm
        c.number_format = MULT if ("P/E" in h or "P/BV" in h or "TEV" in h) else NUM1
LAST = r0 + len(U) - 1
sp_col = colmap["Special Sit"]
wu.conditional_formatting.add(f"A{r0}:{L(len(all_hdr))}{LAST}", FormulaRule(formula=[f"${sp_col}{r0}=\"Y\""], fill=fill_sp))
wu.freeze_panes = f"C{r0}"
wu.auto_filter.ref = f"A{HR}:{L(len(all_hdr))}{LAST}"
widths = [8, 18, 9, 14, 14, 18, 22, 14, 7, 12, 40, 34] + [10] * (len(cols) - 12) + [11] * len(formula_cols) + [10] * len(ciq_cols)
setw(wu, widths)
wu.row_dimensions[HR].height = 42
wu[f"{colmap['Event Note (K, verify)']}{HR}"].comment = Comment("모델 지식 기반 태그(2026년 중반 이전 정보). DART/CIQ로 검증 필요.", "Claude")
wu[f"{colmap['ADV min (억)']}{HR}"].comment = Comment("min(20D,60D) 평균 거래대금, 거래정지일 제외. 억원.", "Claude")
wu[f"{colmap['Mcap 1M avg (억)']}{HR}"].comment = Comment("최근 20영업일 평균 시가총액 (BBAS 참고: 1개월 평균 시총 적용 가능).", "Claude")

# ======================= Summary =======================
wsu = wb.create_sheet("Summary", 1)
wsu.cell(row=1, column=1, value=f"Summary — Korea universe under BBAS liquidity limits (as of {ASOF})").font = f_title
wsu.cell(row=2, column=1, value="모든 수치는 Universe 시트에서 COUNTIFS/AVERAGEIFS로 계산. CIQ 열 평균은 Capital IQ 연결 후 표시(미연결 시 n/a).").font = f_norm
UR = f"Universe!${{c}}${r0}:${{c}}${LAST}"
def rng(h):
    return UR.format(c=colmap[h])
tiers = list(tier_order.keys())
hdrs = ["Tier (Alpha 5d @Book1)", "# names", "% of universe", "Avg Mcap (억)", "Avg ADV20 (억)", "Avg daily turnover", "Avg Vol 60D", "Avg Ret YTD", "Avg Ret 1M", "# Special Sit",
        "# 6% OK Core(3d) @Book1", "# 6% OK Alpha(5d) @Book1", "# 6% OK Alpha(5d) @Book2", "Avg P/E LTM", "Avg P/BV", "Avg TEV/EBITDA", "Avg ROE %", "Avg ROIC %", "Avg Rev g %", "Avg EPS g %", "Avg Div yld %"]
R = 4
for j, h in enumerate(hdrs, 1):
    wsu.cell(row=R, column=j, value=h)
style_header(wsu, R, len(hdrs))
def avgifs(h, crit_rng, crit, extra=""):
    return f'=IFERROR(AVERAGEIFS({rng(h)},{crit_rng},{crit}{extra}),"n/a")'
for i, t in enumerate(tiers + ["TOTAL"], R + 1):
    wsu.cell(row=i, column=1, value=t).font = f_bold if t == "TOTAL" else f_norm
    if t == "TOTAL":
        crit_rng, crit = rng("Tier"), '"*"'
    else:
        crit_rng, crit = rng("Tier"), f'"{t}"'
    fx = [
        f"=COUNTIFS({crit_rng},{crit})",
        f"=B{i}/$B${R + 1 + len(tiers)}",
        avgifs("Mcap (억)", crit_rng, crit), avgifs("ADV20 (억)", crit_rng, crit), avgifs("Daily turnover", crit_rng, crit),
        avgifs("Vol 60D ann.", crit_rng, crit), avgifs("Ret YTD", crit_rng, crit), avgifs("Ret 1M", crit_rng, crit),
        f'=COUNTIFS({crit_rng},{crit},{rng("Special Sit")},"Y")',
        f'=COUNTIFS({crit_rng},{crit},{rng("MaxPos %NAV Book1")},">="&{PPOS})',
        f'=COUNTIFS({crit_rng},{crit},{rng("MaxPos Alpha(5d) %NAV Book1")},">="&{PPOS})',
        f'=COUNTIFS({crit_rng},{crit},{rng("MaxPos Alpha(5d) %NAV Book2")},">="&{PPOS})',
        avgifs("P/E LTM", crit_rng, crit), avgifs("P/BV", crit_rng, crit), avgifs("TEV/EBITDA LTM", crit_rng, crit), avgifs("ROE %", crit_rng, crit),
        avgifs("ROIC % (Return on Capital)", crit_rng, crit), avgifs("Rev growth 1Y %", crit_rng, crit), avgifs("EPS growth 1Y %", crit_rng, crit), avgifs("Div yield %", crit_rng, crit),
    ]
    fmts = [NUM0, PCT, NUM0, NUM1, "0.00%", PCT, PCT, PCT, NUM0, NUM0, NUM0, NUM0, MULT, MULT, MULT, NUM1, NUM1, NUM1, NUM1, NUM1]
    for j, (f, fm) in enumerate(zip(fx, fmts), 2):
        c = wsu.cell(row=i, column=j, value=f); c.font = f_norm; c.number_format = fm; c.border = border
R2 = R + len(tiers) + 4
wsu.cell(row=R2 - 1, column=1, value="By market-cap bucket (1M avg) — 전체 상장종목 기준 (Universe 밖 포함)").font = f_bold
for j, h in enumerate(["Bucket", "# all listed", "# in Universe", "# <3000억 (BBAS ⑥)", "# Excluded (KONEX/SPAC/관리/정지/illiquid)"], 1):
    wsu.cell(row=R2, column=j, value=h)
style_header(wsu, R2, 5)
allc = u.groupby("Mcap_Bucket").size(); inu = u[u.In_Universe].groupby("Mcap_Bucket").size(); evc = u[u.Tier.str.startswith("C")].groupby("Mcap_Bucket").size(); exc = u[u.Tier.str.startswith("X")].groupby("Mcap_Bucket").size()
for i, b in enumerate(["A ≥7000억", "B 3000-7000억", "C <3000억"], R2 + 1):
    for j, v in enumerate([b, int(allc.get(b, 0)), int(inu.get(b, 0)), int(evc.get(b, 0)), int(exc.get(b, 0))], 1):
        c = wsu.cell(row=i, column=j, value=v); c.font = f_norm; c.border = border
        if j > 1: c.number_format = NUM0
i = R2 + 4
for j, v in enumerate(["TOTAL", f"=SUM(B{R2+1}:B{R2+3})", f"=SUM(C{R2+1}:C{R2+3})", f"=SUM(D{R2+1}:D{R2+3})", f"=SUM(E{R2+1}:E{R2+3})"], 1):
    c = wsu.cell(row=i, column=j, value=v); c.font = f_bold; c.border = border; c.number_format = NUM0
wsu.cell(row=i + 1, column=1, value="(정적 값: build 시점 계산. 데이터 갱신 시 스크립트 재실행)").font = Font(name=FONT, size=8, italic=True)
R3 = i + 4
wsu.cell(row=R3 - 1, column=1, value="By market").font = f_bold
for j, h in enumerate(["Market", "# in Universe", "# Special Sit", "Avg Mcap (억)", "Avg ADV20 (억)"], 1):
    wsu.cell(row=R3, column=j, value=h)
style_header(wsu, R3, 5)
for i, mk in enumerate(["KOSPI", "KOSDAQ", "KOSDAQ GLOBAL"], R3 + 1):
    fx = [mk, f'=COUNTIFS({rng("Market")},"{mk}")', f'=COUNTIFS({rng("Market")},"{mk}",{rng("Special Sit")},"Y")', avgifs("Mcap (억)", rng("Market"), f'"{mk}"'), avgifs("ADV20 (억)", rng("Market"), f'"{mk}"')]
    for j, v in enumerate(fx, 1):
        c = wsu.cell(row=i, column=j, value=v); c.font = f_norm; c.border = border; c.number_format = NUM0 if j in (2, 3, 4) else NUM1
R4 = R3 + 6
wsu.cell(row=R4 - 1, column=1, value="Liquidity capacity check (Params 변경 시 재계산)").font = f_bold
for j, h in enumerate(["Metric", "Book1", "Book2"], 1):
    wsu.cell(row=R4, column=j, value=h)
style_header(wsu, R4, 3)
cap = [
    ("Book (억원)", f"={PB1}", f"={PB2}"),
    ("6% position (억원)", f"={PPOS}*{PB1}", f"={PPOS}*{PB2}"),
    ("Required ADV for 3-day exit (억원)", f"={PPOS}*{PB1}/({PPART}*{PDAYS})", f"={PPOS}*{PB2}/({PPART}*{PDAYS})"),
    ("# names allowing full 6% in 3 days", f'=COUNTIFS({rng("MaxPos %NAV Book1")},">="&{PPOS})', f'=COUNTIFS({rng("MaxPos %NAV Book2")},">="&{PPOS})'),
    ("# names allowing ≥ min position (1%) in 3 days", f'=COUNTIFS({rng("MaxPos %NAV Book1")},">="&{PMIN})', f'=COUNTIFS({rng("MaxPos %NAV Book2")},">="&{PMIN})'),
    ("# names full 6% — Alpha sleeve (5 days)", f'=COUNTIFS({rng("MaxPos Alpha(5d) %NAV Book1")},">="&{PPOS})', f'=COUNTIFS({rng("MaxPos Alpha(5d) %NAV Book2")},">="&{PPOS})'),
    ("# names ≥ 3% — Alpha sleeve (5 days)", f'=COUNTIFS({rng("MaxPos Alpha(5d) %NAV Book1")},">=0.03")', f'=COUNTIFS({rng("MaxPos Alpha(5d) %NAV Book2")},">=0.03")'),
    ("# names ≥ 1% — Alpha sleeve (5 days)", f'=COUNTIFS({rng("MaxPos Alpha(5d) %NAV Book1")},">="&{PMIN})', f'=COUNTIFS({rng("MaxPos Alpha(5d) %NAV Book2")},">="&{PMIN})'),
    ("# names ≥ 1% — Event sleeve (10 days)", f'=COUNTIFS({rng("MaxPos Event(10d) %NAV Book1")},">="&{PMIN})', f'=COUNTIFS({rng("MaxPos Event(10d) %NAV Book2")},">="&{PMIN})'),
    ("# names full 6% — Alpha sleeve at 12M trough ADV", f'=COUNTIFS({rng("Trough MaxPos Alpha(5d) %NAV Book1")},">="&{PPOS})', ""),
    ("Sum of MaxPos Alpha(5d) %NAV (gross capacity)", f'=SUM({rng("MaxPos Alpha(5d) %NAV Book1")})', f'=SUM({rng("MaxPos Alpha(5d) %NAV Book2")})'),
    ("# names full 6% under stress ADV (Params haircut)", f'=COUNTIFS({rng("Stress MaxPos %NAV Book1")},">="&{PPOS})', ""),
    ("# names full 6% at 12M trough ADV (empirical)", f'=COUNTIFS({rng("Trough MaxPos %NAV Book1")},">="&{PPOS})', f'=COUNTIFS({rng("Trough MaxPos %NAV Book2")},">="&{PPOS})'),
    ("Median ADV stability (12M trough / current)", f'=MEDIAN({rng("ADV stability (trough/current)")})', ""),
    ("Sum of MaxPos %NAV (gross capacity, 3-day)", f'=SUM({rng("MaxPos %NAV Book1")})', f'=SUM({rng("MaxPos %NAV Book2")})'),
]
for i, (k, a, b) in enumerate(cap, R4 + 1):
    wsu.cell(row=i, column=1, value=k).font = f_norm
    for j, v in enumerate([a, b], 2):
        c = wsu.cell(row=i, column=j, value=v); c.font = f_norm; c.border = border; c.number_format = NUM0 if "%" not in k else PCT
    if "Sum of" in k:
        for j in (2, 3): wsu.cell(row=i, column=j).number_format = "0%"
    if "Alpha sleeve" in k and "Sum" not in k:
        for j in (2, 3): wsu.cell(row=i, column=j).number_format = NUM0
    if "Median ADV" in k:
        wsu.cell(row=i, column=2).number_format = "0.00x"
setw(wsu, [40, 11, 12, 13, 13, 13, 11, 11, 11, 11, 12, 12, 12, 11, 10, 12, 10, 10, 10, 10, 10])
wsu.row_dimensions[R].height = 40

# ======================= Special_Situations =======================
wss = wb.create_sheet("Special_Situations", 2)
S = U[U["SpecialSit"] == "Y"].copy()
S["_e"] = (S["Event_Type"] == "").astype(int)
S = S.sort_values(["_e", "_t", "Mcap"], ascending=[True, True, False])
sd = pd.DataFrame({
    "Code": S["Code"], "Name": S["Name"], "Market": S["Market"], "Tier": S["Tier"], "Event Type (K)": S["Event_Type"], "Event Note (K, verify)": S["Event_Note"],
    "Quant Flags (data)": S["Quant_Flags"], "Mcap (억)": S["Mcap"] / E, "ADV20 (억)": S["ADV20"] / E, "MaxPos %NAV @500억": S["MaxPosPct_500"], "MaxPos %NAV @1000억": S["MaxPosPct_1000"],
    "DTL 6% @500억 (d)": S["DTL6_500"], "MaxPos Alpha(5d) @500억": S["MaxPos_500_alpha"], "Short cap": S["Short_Cap"], "Ret 1M": S["Ret_1M"], "Ret YTD": S["Ret_YTD"], "% from 52wH": S["Pct_from_52wH"], "Vol spike": S["VolSpike"], "±15% days 40D": S["Vol15_40D"],
})
wss.cell(row=1, column=1, value=f"Special Situations in Universe — {len(sd)} names ({(S['Event_Type']!='').sum()} with event tag, rest = quant signal only). Event tag = 모델 지식(검증 필요), Quant = 데이터.").font = f_bold
write_df(wss, sd, 3, fmts={"Mcap (억)": NUM0, "ADV20 (억)": NUM1, "MaxPos %NAV @500억": PCT, "MaxPos %NAV @1000억": PCT, "DTL 6% @500억 (d)": NUM1, "MaxPos Alpha(5d) @500억": PCT, "Short cap": PCT, "Ret 1M": PCT, "Ret YTD": PCT, "% from 52wH": PCT, "Vol spike": MULT, "±15% days 40D": NUM0})
setw(wss, [8, 18, 9, 22, 14, 50, 40, 10, 10, 10, 10, 10, 8, 8, 8, 9, 8, 8])
wss.freeze_panes = "C4"; wss.auto_filter.ref = f"A3:{L(len(sd.columns))}{3+len(sd)}"
wss.row_dimensions[3].height = 36

# ======================= Pref_Pairs =======================
wpp = wb.create_sheet("Pref_Pairs")
pp = prefs[(prefs["PrefADV20"] >= 5 * E)].copy()
pd_ = pd.DataFrame({
    "Pref": pp["Pref"], "Pref Code": pp["PrefCode"], "Common": pp["Common"], "Common Code": pp["CommonCode"], "Pref Px": pp["PrefPx"], "Common Px": pp["ComPx"],
    "Discount (Pref/Common-1)": pp["Discount"], "Pref Mcap (억)": pp["PrefMcap"] / E, "Pref ADV20 (억)": pp["PrefADV20"] / E, "Common ADV20 (억)": pp["ComADV20"] / E,
    "Pref Tier": pp["PrefTier"], "Common Tier": pp["ComTier"], "Pref MaxPos @500억": pp["PrefMaxPos500"], "Common MaxPos @500억": pp["ComMaxPos500"],
})
wpp.cell(row=1, column=1, value=f"Preferred vs Common — {len(pd_)} pairs with pref ADV20 ≥ 5억 (as of {ASOF}). 전체 {len(prefs)} 쌍 중. Pair 전략은 BBAS: 양 leg 비중 차이 4% 이내").font = f_bold
end = write_df(wpp, pd_, 3, fmts={"Pref Px": NUM0, "Common Px": NUM0, "Discount (Pref/Common-1)": PCT, "Pref Mcap (억)": NUM0, "Pref ADV20 (억)": NUM1, "Common ADV20 (억)": NUM1, "Pref MaxPos @500억": PCT, "Common MaxPos @500억": PCT})
wpp.cell(row=end + 2, column=1, value="Target discount (input, 가정)").font = f_norm
c = wpp.cell(row=end + 2, column=2, value=-0.45); c.font = f_blue; c.fill = fill_yel; c.number_format = PCT
wpp.cell(row=end + 3, column=1, value="→ Pref upside if discount narrows to target (보통주 불변 가정)").font = f_norm
for i in range(4, 4 + len(pd_)):
    c = wpp.cell(row=i, column=15, value=f"=(1+$B${end+2})/(1+G{i})-1"); c.number_format = PCT; c.font = f_norm; c.border = border
wpp.cell(row=3, column=15, value="Upside to target disc."); style_header(wpp, 3, 15)
setw(wpp, [18, 9, 14, 9, 10, 10, 12, 10, 10, 10, 22, 22, 10, 10, 12])
wpp.freeze_panes = "C4"

# ======================= Holdco_NAV =======================
wh = wb.create_sheet("Holdco_NAV")
wh.cell(row=1, column=1, value="Holdco listed-stake coverage (상장자회사 지분가치 / 지주 시총). 지분율은 모델 기억 기반 근사치(파란색) → CIQ/DART 검증 필수. 순차입금·비상장 자산 미반영.").font = f_bold
holdcos = [
    ("402340", "SK스퀘어", [("000660", "SK하이닉스", 0.201)]),
    ("028260", "삼성물산", [("005930", "삼성전자", 0.0501), ("207940", "삼성바이오로직스", 0.4306), ("0126Z0", "삼성에피스홀딩스", 0.4306), ("032830", "삼성생명", 0.1934), ("018260", "삼성에스디에스", 0.1708)]),
    ("000150", "두산", [("034020", "두산에너빌리티", 0.304), ("454910", "두산로보틱스", 0.682), ("131970", "두산테스나", 0.0)]),
    ("003550", "LG", [("051910", "LG화학", 0.333), ("066570", "LG전자", 0.337), ("051900", "LG생활건강", 0.340), ("032640", "LG유플러스", 0.377), ("064400", "LG씨엔에스", 0.4995)]),
    ("034730", "SK", [("096770", "SK이노베이션", 0.559), ("017670", "SK텔레콤", 0.306), ("402340", "SK스퀘어", 0.306), ("011790", "SKC", 0.406), ("001740", "SK네트웍스", 0.439), ("326030", "SK바이오팜", 0.64)]),
    ("267250", "HD현대", [("009540", "HD한국조선해양", 0.37), ("267260", "HD현대일렉트릭", 0.37), ("443060", "HD현대마린솔루션", 0.62), ("267270", "HD건설기계", 0.30)]),
    ("180640", "한진칼", [("003490", "대한항공", 0.261)]),
    ("000670", "영풍", [("010130", "고려아연", 0.254)]),
    ("001040", "CJ", [("097950", "CJ제일제당", 0.446), ("035760", "CJ ENM", 0.421)]),
    ("004990", "롯데지주", [("011170", "롯데케미칼", 0.253), ("023530", "롯데쇼핑", 0.40), ("005300", "롯데칠성", 0.45), ("280360", "롯데웰푸드", 0.48)]),
    ("002790", "아모레퍼시픽홀딩스", [("090430", "아모레퍼시픽", 0.372)]),
    ("001800", "오리온홀딩스", [("271560", "오리온", 0.374)]),
    ("008930", "한미사이언스", [("128940", "한미약품", 0.414)]),
    ("005440", "현대지에프홀딩스", [("069960", "현대백화점", 0.37), ("453340", "현대그린푸드", 0.40)]),
    ("004800", "효성", [("298040", "효성중공업", 0.325), ("298020", "효성티앤씨", 0.203)]),
]
mc = u.set_index("Code")
hh = ["Holdco Code", "Holdco", "Holdco Mcap (억)", "Holdco Tier", "Sub Code", "Subsidiary", "Stake % (input)", "Sub Mcap (억)", "Stake value (억)", "Holdco listed-stake NAV (억)", "Coverage = NAV/Mcap", "Implied discount = 1 - Mcap/NAV"]
for j, h in enumerate(hh, 1):
    wh.cell(row=3, column=j, value=h)
style_header(wh, 3, len(hh))
r = 4
for hc, hn, subs in holdcos:
    if hc not in mc.index:
        continue
    first = r
    for sc, sn, st in subs:
        if sc not in mc.index:
            continue
        vals = [hc, hn, mc.loc[hc, "Mcap"] / E, mc.loc[hc, "Tier"], sc, sn, st, mc.loc[sc, "Mcap"] / E]
        for j, v in enumerate(vals, 1):
            c = wh.cell(row=r, column=j, value=v); c.font = f_blue if j == 7 else f_norm; c.border = border
            c.number_format = PCT if j == 7 else (NUM0 if j in (3, 8) else "General")
            if j == 7: c.fill = fill_yel
        c = wh.cell(row=r, column=9, value=f"=G{r}*H{r}"); c.number_format = NUM0; c.font = f_norm; c.border = border
        r += 1
    last = r - 1
    for rr in range(first, last + 1):
        c = wh.cell(row=rr, column=10, value=f"=SUM($I${first}:$I${last})"); c.number_format = NUM0; c.font = f_norm; c.border = border
        c = wh.cell(row=rr, column=11, value=f"=J{rr}/C{rr}"); c.number_format = "0.00x"; c.font = f_norm; c.border = border
        c = wh.cell(row=rr, column=12, value=f"=1-C{rr}/J{rr}"); c.number_format = PCT; c.font = f_norm; c.border = border
    r += 1
wh.cell(row=r + 1, column=1, value="해석: Coverage > 1.0x = 상장자회사 지분가치만으로 지주 시총 초과(비상장 자산·순차입금 제외). 지분율 0 = 미확인. Mcap은 Universe build 시점 값(정적).").font = Font(name=FONT, size=8, italic=True)
setw(wh, [10, 16, 12, 22, 9, 16, 10, 12, 12, 14, 12, 14])
wh.freeze_panes = "A4"

# ======================= Ideas =======================
wi = wb.create_sheet("Ideas")
def g(code, key):
    return mc.loc[code, key] if code in mc.index else np.nan
def disc(pc, cc):
    return g(pc, "Close") / g(cc, "Close") - 1
ideas = [
    dict(n=1, idea="삼성전기우 / 삼성전기 pref-common pair", dir="Long pref / Short common", legs="009155 / 009150", typ="Pref discount",
         thesis=f"괴리율 {disc('009155','009150'):.0%}. 양 leg 모두 A1(우선주 ADV20 {g('009155','ADV20')/E:.0f}억). 2차 상법개정·자사주 소각 의무화 논의는 우선주 배당·청산가치 재평가 촉매. 보통주 YTD {g('009150','Ret_YTD'):+.0%}로 모멘텀 과열 → 보통주 short leg가 hedge 겸 alpha.",
         cat="자사주 소각 의무화 입법, 배당 확대, 우선주 ETF/패시브 수급", size500="6% / 6% (pair 차이 4% 이내)", up="괴리율 -63% → -50% 수렴 시 pref 상대수익 +35%; -55%면 +22%", dn="괴리율 -70%로 확대 시 -19%. 보통주 급등 국면에서 pref 소외 지속",
         risk="괴리율 mean-reversion 시점 불확실. 대차 가능 여부/비용", da="괴리율은 수년간 -50~-65% 박스. 촉매 없이는 carry(배당차)만 남음. 2025-26 랠리에서 pref 소외 → 구조적일 수 있음", need="배당 이력, 우선주 발행주식수·유통물량, 대차잔고"),
    dict(n=2, idea="현대차2우B / 현대차 pair", dir="Long pref / Short common", legs="005387 / 005380", typ="Pref discount",
         thesis=f"괴리율 {disc('005387','005380'):.0%}, 2우B ADV20 {g('005387','ADV20')/E:.0f}억으로 1,000억 Book 6%도 3일 청산 가능. 현대차 밸류업(TSR 35%)·자사주 소각은 우선주에 동일 적용. 지배구조 개편 시 우선주 매입·소각 시나리오.",
         cat="자사주 소각, 지배구조 개편(모비스 순환출자), 배당 확대", size500="6% / 6%", up="-52% → -40% 수렴 시 +25%", dn="-58% 확대 시 -13%",
         risk="관세·자동차 사이클은 pair로 중립. 우선주 유동성은 보통주 대비 1/10", da="현대차 우선주 괴리율은 이미 2021년 이후 좁혀진 상태. 추가 수렴 촉매 제한적", need="배당 정책, 자사주 소각 대상에 우선주 포함 여부"),
    dict(n=3, idea="SK스퀘어 NAV discount vs SK하이닉스", dir="Long SK스퀘어 / Short SK하이닉스 (지분율 만큼)", legs="402340 / 000660", typ="Holdco NAV",
         thesis=f"SK스퀘어 시총 {g('402340','Mcap')/E/1e4:.0f}조 vs 하이닉스 20.1% 지분가치 {g('000660','Mcap')*0.201/E/1e4:.0f}조 → 커버리지 {g('000660','Mcap')*0.201/g('402340','Mcap'):.2f}x (상장지분만). 자사주 소각·주주환원 확대 시 디스카운트 축소. 양 leg A1.",
         cat="자사주 매입/소각, 비핵심 자회사 매각, 상법개정(지주 디스카운트 축소)", size500="6% / 6% (베타 조정 시 hedge ratio ≈ 1.0~1.2)", up="디스카운트 45% → 35%: 상대수익 +18%", dn="디스카운트 55%로 확대: -18%. 하이닉스 급등 시 스퀘어 beta<1로 lag",
         risk="하이닉스 변동성(±15%일 4회 → 스퀘어 Short cap -4% 해당) → 사이징 제약. 순차입금·비상장 자산(11번가 등) 미반영", da="지주 디스카운트는 한국 시장 구조적(세금·지배주주). 하이닉스 랠리 중 '스퀘어 = 저베타 하이닉스'로 lag 지속 가능", need="스퀘어 순차입금, 자사주 비중, 비상장 자회사 가치"),
    dict(n=4, idea="삼성물산 listed-stake coverage", dir="Long 삼성물산 / Short 삼성전자·삼성바이오 basket", legs="028260 / 005930, 207940, 0126Z0", typ="Holdco NAV",
         thesis="Holdco_NAV 시트: 삼성전자 5.0%·바이오로직스/에피스홀딩스 43%·삼성생명 19%·SDS 17% 지분가치가 시총 대비 커버리지 >1.5x(추정). 삼성 지배구조 개편·자사주 소각·보험업법 이슈 모두 촉매.",
         cat="자사주 소각, 지배구조 개편, 삼성생명 보험업법", size500="6% / basket 6%", up="디스카운트 10%p 축소 시 +15~20%", dn="개편 지연·건설 부문 부진 시 -10~15%",
         risk="상법·세제 변경 속도. 삼성전자 랠리 시 물산 lag", da="삼성물산 디스카운트는 2015년 합병 이후 구조적. 촉매가 '논의'에 머무름", need="순차입금, 비상장 자산(바이오 외), 자사주 비율"),
    dict(n=5, idea="삼성에피스홀딩스 post-spin", dir="Long (단독) 또는 Long 에피스홀딩스 / Short 삼성바이오로직스", legs="0126Z0 / 207940", typ="Spin-off",
         thesis=f"2025-11 인적분할 신설법인. A1 유동성(ADV20 {g('0126Z0','ADV20')/E:.0f}억). 분할 후 지수편입·패시브 수급 정상화 구간이 통상 6~12개월. 바이오시밀러 pure-play 재평가.",
         cat="지수 편입, 신제품 승인, 커버리지 개시", size500="6% / 6%", up="peer 멀티플 적용 시 +20~30% (CIQ 데이터 필요)", dn="시밀러 가격경쟁 심화 시 -20%",
         risk="밸류에이션 데이터 부재 → 확인 전 사이징 불가", da="post-spin 아노말리는 대형 spin에서 약함. 이미 재상장 10개월 경과", need="EV/EBITDA, 파이프라인, 순현금"),
    dict(n=6, idea="한화 분할 재상장 + 한화머시너리앤서비스홀딩스", dir="Monitor → Long 신설법인 (수급 왜곡 해소) / Short 한화(존속)", legs="0220W0 / 000880", typ="Spin-off / 재상장",
         thesis=f"2026-08 재상장 8영업일 경과. 신설법인 ADV20 {g('0220W0','ADV20')/E:.0f}억(B tier, 시총 {g('0220W0','Mcap')/E:.0f}억). 분할 직후 지수/패시브 강제매도 구간. Event Play(10%, 40거래일) 사전승인 대상.",
         cat="지수 편입/제외 확정, 승계 구조 공시, 첫 실적", size500="Event Play ≤10% / 40거래일", up="강제매도 해소 시 +15~25%", dn="승계용 저평가 유지 시 -15%",
         risk="분할 목적(승계) 상 신설법인 저평가가 의도적일 수 있음", da="한화 그룹 분할은 승계 최적화가 목적 → 소액주주 가치 제고와 이해상충. 시총 3천억~7천억 bucket 7% 제약", need="분할 비율, 자산 구성, 대주주 지분"),
    dict(n=7, idea="롯데렌탈 PE 인수 후 잔여지분 이벤트", dir="Event arb (Long) — 조건부", legs="089860", typ="Tender / 상폐",
         thesis=f"어피너티 인수(2025) 후 YTD {g('089860','Ret_YTD'):+.0%}, 거래대금 급증 {g('089860','VolSpike'):.1f}x. 잔여지분 공개매수·상폐 추진 시 spread 거래. A3(ADV20 {g('089860','ADV20')/E:.0f}억) → 500억 Book 2~6%.",
         cat="공개매수 공고, 상폐 요건(95%) 충족 여부", size500=f"2~3% (DTL 6% = {g('089860', 'DTL6_500'):.1f}일)", up="공개매수 프리미엄 10~20% (가격 미확인)", dn="공개매수 무산/저가 시 -20~30%",
         risk="공개매수 가격·시점 미확인 → 현재는 monitor. 유동성 낮음", da="이미 +68% 반영. 잔여 upside는 프리미엄 협상에 의존", need="DART 공개매수 공고, 대주주 지분율"),
    dict(n=8, idea="HD건설기계 post-merger", dir="Long", legs="267270", typ="Merger integration",
         thesis=f"HD현대건설기계+인프라코어 합병(2026-01 완료). A1(ADV20 {g('267270','ADV20')/E:.0f}억), YTD {g('267270','Ret_YTD'):+.0%}. 중복 비용 제거·구매 통합 시너지가 합병 후 2~3분기에 실적으로 확인되는 패턴.",
         cat="시너지 가이던스, 분기 실적, 신흥국 인프라 수요", size500="4~6%", up="시너지 반영 시 +20~30%", dn="건설장비 downcycle 시 -20%",
         risk="합병 회계(PPA) 노이즈로 밸류 비교 어려움", da="건설기계는 사이클 산업. 시너지보다 중국·신흥국 수요가 주가 결정", need="합병 후 EV/EBITDA, 시너지 가이던스"),
    dict(n=9, idea="밸류업/상법개정 저PBR 지주 basket vs KOSPI200", dir="Long basket / Short K200 futures", legs="028260, 034730, 003550, 000150, 078930, 001040 + 금융지주", typ="Governance basket",
         thesis="2차 상법개정(집중투표·감사위원 분리선출), 자사주 소각 의무화 논의, 배당소득 분리과세 → 저PBR·자사주 多 지주/금융지주 구조적 수혜. 모든 종목 A1 유동성. 선물 hedge로 Net 15% 한도·Gross-cut 3일 청산 대응.",
         cat="입법 일정, 자사주 소각 공시, 밸류업 공시", size500="종목당 2~3%, basket 15~20% / 선물 15~20%", up="PBR 0.5x→0.65x 리레이팅 시 +25~30%", dn="입법 지연·시장 조정 시 -10% (hedge 후)",
         risk="2025-26 랠리로 이미 일부 반영(삼성물산 YTD +52%, SK +110%). 팩터 crowding", da="저PBR 리레이팅은 2024 밸류업 이후 3년째 테마 → 신선도 하락. 실제 소각 없이 '계획'만 반복", need="PBR, 자사주 비율, 배당성향 (CIQ)"),
    dict(n=10, idea="과열 테마 Short screen (건설·전선·로봇)", dir="Short candidates (사이징 제약)", legs="002990, 047040, 006340, 000500, 277810", typ="Momentum reversal",
         thesis=f"금호건설 YTD {g('002990','Ret_YTD'):+.0%}(±15%일 7회), 대우건설 YTD {g('047040','Ret_YTD'):+.0%}, 대원전선·가온전선 ±15%일 5회. BBAS 룰상 ±15% 2회 이상 → Short cap -4%. 40거래일 변동성 감소 후 진입.",
         cat="거래대금 감소, 유상증자/대주주 매도 공시, 실적 부진", size500="-2~-4% (Short cap -4%)", up="고점 대비 -30~50% 되돌림", dn="테마 연장 시 -30% (일 ±30% 상한가 리스크)",
         risk="상한가 연속 리스크, 대차 부족·비용, VaR 이탈(1D 1VaR) → Gross-cut 트리거", da="테마 과열 short는 타이밍 실패 시 BBAS Book-cut(-6% DD)의 주범. 개별 short보다 섹터 ETF/선물 hedge가 룰 친화적", need="대차잔고·대차비용, 공매도 잔고, 주주구성"),
]
ih = ["#", "Idea", "Direction", "Legs (codes)", "Type", "Thesis (data-backed)", "Catalyst", "Sizing @500억 (BBAS)", "Upside (가정)", "Downside (가정)", "Key risk", "Devil's advocate", "Data needed (CIQ/DART)"]
wi.cell(row=1, column=1, value=f"Investment Ideas — 특수상황/유동성 룰 내 실행 가능. Upside/Downside는 정량 가정(모델링 아님). as of {ASOF}").font = f_bold
for j, h in enumerate(ih, 1):
    wi.cell(row=3, column=j, value=h)
style_header(wi, 3, len(ih))
for i, d in enumerate(ideas, 4):
    vals = [d["n"], d["idea"], d["dir"], d["legs"], d["typ"], d["thesis"], d["cat"], d["size500"], d["up"], d["dn"], d["risk"], d["da"], d["need"]]
    for j, v in enumerate(vals, 1):
        c = wi.cell(row=i, column=j, value=v); c.font = f_norm; c.border = border; c.alignment = Alignment(wrap_text=True, vertical="top")
    wi.row_dimensions[i].height = 120
setw(wi, [4, 26, 22, 18, 14, 60, 30, 20, 26, 26, 30, 40, 26])
wi.freeze_panes = "C4"


# ======================= PreTrade_Check =======================
wt = wb.create_sheet("PreTrade_Check", 3)
wt.cell(row=1, column=1, value="Pre-trade BBAS compliance check — 파란색 입력(Code / Side / Weight %NAV / Sleeve: C=Core·Hedge 3일, A=Alpha 5일, E=Event Play 승인 10일·한도 10%). 다이나믹 한도 = min(룰 한도, 참여율×sleeve DTL×ADV/Book). 샘플 포트는 Ideas 기반 예시이며 Net 한도 위반을 의도적으로 남겨둠(K200 선물 -10% hedge 시 해소).").font = f_bold
wt.cell(row=2, column=1, value="Book = Params B4 (억원). 룰 한도: 일반 ±6%, 시총 1-2위 Long 8%, Event Play 10%, ±15%일 2회+ Short -4%. Liquidity OK = DTL ≤ sleeve 목표일수. Trim = DTL > 목표×Params B24. VaR proxy = 2.33 × 60D 일변동성 (상관 미반영).").font = Font(name=FONT, size=8, italic=True)
th = ["Code", "Side (L/S)", "Weight %NAV", "Sleeve (C/A/E)", "Name", "Tier", "Mcap Bucket", "ADV min (억)", "±15% days", "1D VaR proxy", "Position (억)", "Limit %NAV", "Limit OK", "Sleeve exit capacity (억)", "DTL (days)", "Liquidity OK (DTL ≤ sleeve target)", "Trough DTL (days)", "VaR contrib (|w|×VaR)", "Long w", "Short w", "Top5 Long", "Top5 Short", "Gross-cut day-1 sell (억)", "Day-1 % of ADV", "Participation OK", "Sleeve DTL target", "Dynamic limit %NAV (min(rule, liquidity))", "Dynamic OK", "Trim trigger (DTL > target×k)", "(Σw)i (cov·w)", "Marginal VaR (cov)", "Component VaR (cov)", "% of portfolio VaR", "Standalone − Component (hedge benefit)"]
TR = 4
for j, h in enumerate(th, 1):
    wt.cell(row=TR, column=j, value=h)
style_header(wt, TR, len(th))
sample = [("009155", "L", 0.06, "A"), ("009150", "S", -0.04, "C"), ("005387", "L", 0.05, "A"), ("005380", "S", -0.05, "C"), ("402340", "L", 0.04, "C"), ("000660", "S", -0.04, "C"),
          ("028260", "L", 0.04, "C"), ("005930", "S", -0.02, "C"), ("207940", "S", -0.02, "C"), ("0126Z0", "L", 0.03, "A"), ("267270", "L", 0.03, "A"), ("000150", "L", 0.02, "A"),
          ("003550", "L", 0.02, "A"), ("034730", "L", 0.02, "C"), ("105560", "L", 0.02, "C"), ("0220W0", "L", 0.02, "E"), ("089860", "L", 0.02, "E"), ("002990", "S", -0.02, "A"), ("006340", "S", -0.02, "A"), ("241560", "L", 0.02, "A"),
          ("026960", "L", 0.01, "E"), ("012630", "L", 0.01, "E")]
NROWS = 40
r1, rN = TR + 1, TR + NROWS
pr = rN + 2
pr_var = pr + 20   # row of "Portfolio sigma (cov)" in the checks table (index 19 below +1)
UA = f"Universe!$A${r0}:$A${LAST}"
def ucol(h):
    return f"Universe!${colmap[h]}${r0}:${colmap[h]}${LAST}"
GC, T1, T2, VARL = "Params!$B$17", "Params!$B$18", "Params!$B$19", "Params!$B$20"
for i in range(NROWS):
    row = r1 + i
    code, side, w, ev = sample[i] if i < len(sample) else ("", "", None, "")
    for j, v in enumerate([code, side, w, ev], 1):
        c = wt.cell(row=row, column=j, value=v); c.font = f_blue; c.fill = fill_yel; c.border = border
        if j == 3: c.number_format = PCT
    A, B, C, Dd = f"$A{row}", f"$B{row}", f"$C{row}", f"$D{row}"
    m = f"MATCH({A},{UA},0)"
    look = lambda h, fmt="General": f'=IF({A}="","",IFERROR(INDEX({ucol(h)},{m}),"n/a"))'
    fx = [
        (look("Name"), "General"), (look("Tier"), "General"), (look("Mcap Bucket"), "General"), (look("ADV min (억)"), NUM1), (look("±15% days 40D"), NUM0), (look("1D 99% VaR proxy"), PCT),
        (f'=IF({C}="","",ABS({C})*{PB1})', NUM1),
        (f'=IF({C}="","",IF({Dd}="E",Params!$B$10,IF({B}="L",IF(OR({A}={T1},{A}={T2}),0.08,{PPOS}),IF(N($I{row})>=2,0.04,{PPOS}))))', PCT),
        (f'=IF({C}="","",IF(ABS({C})<=$L{row}+1E-9,"OK","BREACH"))', "General"),
        (f'=IF(OR({C}="",NOT(ISNUMBER($H{row}))),"",{PPART}*$Z{row}*$H{row})', NUM1),
        (f'=IF(OR({C}="",NOT(ISNUMBER($H{row}))),"",IF($H{row}>0,$K{row}/({PPART}*$H{row}),"n/a"))', NUM1),
        (f'=IF($O{row}="","",IF(AND(ISNUMBER($O{row}),$O{row}<=$Z{row}+1E-9),"OK","SLOW"))', "General"),
        (f'=IF(OR({C}="",NOT(ISNUMBER($H{row}))),"",IFERROR($K{row}/({PPART}*INDEX({ucol("ADV20 12M trough (억)")},{m})),"n/a"))', NUM1),
        (f'=IF(OR({C}="",NOT(ISNUMBER($J{row}))),"",ABS({C})*$J{row})', PCT),
        (f'=IF({C}="","",MAX({C},0))', PCT), (f'=IF({C}="","",MAX(-{C},0))', PCT),
        (f'=IF({C}="","",IF(AND($S{row}>0,COUNTIF($S${r1}:$S${rN},">"&$S{row})<5),$S{row},0))', PCT),
        (f'=IF({C}="","",IF(AND($T{row}>0,COUNTIF($T${r1}:$T${rN},">"&$T{row})<5),$T{row},0))', PCT),
        (f'=IF({C}="","",$K{row}*{GC}*IF(OR($U{row}>0,$V{row}>0),2/3,1)/{PDAYS})', NUM1),
        (f'=IF(OR({C}="",NOT(ISNUMBER($H{row}))),"",IF($H{row}>0,$W{row}/$H{row},"n/a"))', PCT),
        (f'=IF($X{row}="","",IF(AND(ISNUMBER($X{row}),$X{row}<={PPART}+1E-9),"OK","EXCEEDS"))', "General"),
        (f'=IF({C}="","",IF({Dd}="E",{PDE},IF({Dd}="A",{PDA},{PDAYS})))', NUM0),
        (f'=IF(OR({C}="",NOT(ISNUMBER($H{row}))),"",MIN($L{row},{PPART}*$Z{row}*$H{row}/{PB1}))', PCT),
        (f'=IF($AA{row}="","",IF(ABS({C})<=$AA{row}+1E-9,"OK","OVER"))', "General"),
        (f'=IF(OR({C}="",NOT(ISNUMBER($O{row}))),"",IF($O{row}>$Z{row}*{PTRIM},"TRIM","-"))', "General"),
        (f'=IF({C}="","",IFERROR(SUMPRODUCT(INDEX({COV_MAT},MATCH({A},{COV_CODES},0),0),{COV_W}),"n/a"))', "0.000000"),
        (f'=IF(OR($AD{row}="",NOT(ISNUMBER($AD{row}))),"",IF($B${pr_var}>0,{PZ}*$AD{row}/$B${pr_var},0))', PCT),
        (f'=IF(OR($AE{row}="",NOT(ISNUMBER($AE{row}))),"",{C}*$AE{row})', "0.00%"),
        (f'=IF(OR($AF{row}="",NOT(ISNUMBER($AF{row}))),"",IF($B${pr_var+1}<>0,$AF{row}/ABS($B${pr_var+1}),0))', PCT),
        (f'=IF(OR($AF{row}="",NOT(ISNUMBER($AF{row}))),"",$R{row}-$AF{row})', "0.00%"),
    ]
    for k, (f, fmt) in enumerate(fx, 5):
        c = wt.cell(row=row, column=k, value=f); c.font = f_norm; c.border = border; c.number_format = fmt
# portfolio-level checks
wt.cell(row=pr, column=1, value="Portfolio-level checks").font = f_bold
checks = [
    ("Gross exposure", f"=SUMPRODUCT(ABS($C${r1}:$C${rN}))", "PM-specific Gross Cap (BBAS 확인)", "", "0%"),
    ("Net exposure (dollar)", f"=SUM($C${r1}:$C${rN})", "Min(Gross×10%, 15%)", f"=MIN(0.1*B{pr+1},0.15)", PCT),
    ("Net OK", f'=IF(ABS(B{pr+2})<=D{pr+2}+1E-9,"OK","BREACH — Short hedge/선물 추가 필요")', "", "", "General"),
    ("Long TOP5 sum", f"=SUM($U${r1}:$U${rN})", "≤ 25%", 0.25, PCT),
    ("Short TOP5 sum", f"=SUM($V${r1}:$V${rN})", "≤ 25%", 0.25, PCT),
    ("TOP5 OK", f'=IF(AND(B{pr+4}<=D{pr+4}+1E-9,B{pr+5}<=D{pr+5}+1E-9),"OK","BREACH")', "", "", "General"),
    ("Mid-cap (3000-7000억) aggregate |w|", f'=SUMPRODUCT(($G${r1}:$G${rN}="B 3000-7000억")*($D${r1}:$D${rN}<>"E")*ABS(N(+$C${r1}:$C${rN})))', "≤ 7% (Event Play 승인 제외)", "=Params!$B$11", PCT),
    ("Mid-cap OK", f'=IF(B{pr+7}<=D{pr+7}+1E-9,"OK","BREACH")', "", "", "General"),
    ("Sub-3000억 / non-universe |w| (승인 없이)", f'=SUMPRODUCT((LEFT($F${r1}:$F${rN},1)="C")*($D${r1}:$D${rN}<>"E")*ABS(N(+$C${r1}:$C${rN})))+SUMPRODUCT(($F${r1}:$F${rN}="n/a")*($D${r1}:$D${rN}<>"E")*ABS(N(+$C${r1}:$C${rN})))', "0 (BBAS ⑥)", 0, PCT),
    ("Event Play aggregate |w| (sleeve E)", f'=SUMPRODUCT(($D${r1}:$D${rN}="E")*ABS(N(+$C${r1}:$C${rN})))', "제안 bucket 5-10% (미팅 확정)", 0.10, PCT),
    ("Liquid sleeve share of gross (DTL ≤ Core days)", f'=IFERROR(SUMPRODUCT(ISNUMBER($O${r1}:$O${rN})*(N(+$O${r1}:$O${rN})<={PDAYS})*ABS(N(+$C${r1}:$C${rN})))/B{pr+1},"n/a")', "≥ 40% (제안)", f"={PLIQ}", PCT),
    ("# names over dynamic limit", f'=COUNTIF($AB${r1}:$AB${rN},"OVER")', "0", 0, NUM0),
    ("# names on trim trigger", f'=COUNTIF($AC${r1}:$AC${rN},"TRIM")', "모니터링", 0, NUM0),
    ("# single-name limit breaches", f'=COUNTIF($M${r1}:$M${rN},"BREACH")', "0", 0, NUM0),
    ("# names not exitable in 3 days", f'=COUNTIF($P${r1}:$P${rN},"SLOW")', "0 (또는 liquid hedge로 상쇄)", 0, NUM0),
    ("# names exceeding day-1 participation in gross-cut", f'=COUNTIF($Y${r1}:$Y${rN},"EXCEEDS")', "0", 0, NUM0),
    ("Portfolio 1D 99% VaR — undiversified (Σ|w|×VaR)", f"=-SUM($R${r1}:$R${rN})", "vs limit", f"={VARL}", PCT),
    ("Portfolio 1D 99% VaR — zero-correlation (√Σ(w×VaR)²)", f"=-SQRT(SUMPRODUCT($R${r1}:$R${rN},$R${r1}:$R${rN}))", "vs limit (실제는 두 값 사이; pair는 상관 높아 zero-corr 쪽에 가까움)", f"={VARL}", PCT),
    ("Gross-cut scenario: total to sell over 3 days (억)", f"=B{pr+1}*{PB1}*{GC}", "1일 최소 1/3, TOP5는 2/3", f"=B{pr+19}/3", NUM1),
    ("Portfolio sigma (daily, 60D cov)", f"=SQRT(MAX(0,SUMPRODUCT(N(+$C${r1}:$C${rN}),N(+$AD${r1}:$AD${rN}))))", "w'Σw ^0.5 (Cov 시트)", "", "0.000%"),
    ("Portfolio 1D 99% VaR — covariance-based (Σ component)", f"=-{PZ}*B{pr+20}", "vs limit — pair hedge 반영. 실제 BBAS는 60D 히스토리컬 P&L 기준", f"={VARL}", PCT),
    ("VaR OK (cov-based)", f'=IF(B{pr+21}>=D{pr+21},"OK","BREACH — Gross 축소 또는 hedge 강화")', "", "", "General"),
    ("Diversification benefit (standalone sum − cov VaR)", f"=-B{pr+17}-(-B{pr+21})", "hedge/분산 효과", "", PCT),
    ("Gross allowed at VaR limit (linear scale)", f'=IF(B{pr+21}<>0,B{pr+1}*D{pr+21}/B{pr+21},"n/a")', "현재 구조 유지 시 -1% 한도에 맞는 Gross", "", "0%"),
]
for j, h in enumerate(["Check", "Value", "Rule", "Limit", ""], 1):
    wt.cell(row=pr + 0, column=j + 0, value=h if j > 1 else "Portfolio-level checks")
style_header(wt, pr, 4)
for i, (k, f, rule, lim, fmt) in enumerate(checks, pr + 1):
    wt.cell(row=i, column=1, value=k).font = f_norm
    c = wt.cell(row=i, column=2, value=f); c.font = f_norm; c.number_format = fmt; c.border = border
    wt.cell(row=i, column=3, value=rule).font = f_norm
    c = wt.cell(row=i, column=4, value=lim); c.font = f_norm; c.number_format = fmt; c.border = border
wt.conditional_formatting.add(f"M{r1}:M{rN}", FormulaRule(formula=[f'M{r1}="BREACH"'], fill=PatternFill("solid", fgColor="F8CBAD")))
wt.conditional_formatting.add(f"P{r1}:P{rN}", FormulaRule(formula=[f'P{r1}="SLOW"'], fill=PatternFill("solid", fgColor="F8CBAD")))
wt.conditional_formatting.add(f"Y{r1}:Y{rN}", FormulaRule(formula=[f'Y{r1}="EXCEEDS"'], fill=PatternFill("solid", fgColor="F8CBAD")))
wt.conditional_formatting.add(f"B{pr+1}:B{pr+24}", FormulaRule(formula=[f'ISNUMBER(SEARCH("BREACH",B{pr+1}))'], fill=PatternFill("solid", fgColor="F8CBAD")))
wt.conditional_formatting.add(f"AB{r1}:AB{rN}", FormulaRule(formula=[f'AB{r1}="OVER"'], fill=PatternFill("solid", fgColor="F8CBAD")))
wt.conditional_formatting.add(f"AC{r1}:AC{rN}", FormulaRule(formula=[f'AC{r1}="TRIM"'], fill=PatternFill("solid", fgColor="FFE699")))
setw(wt, [9, 8, 9, 8, 16, 20, 13, 10, 8, 9, 10, 9, 9, 11, 8, 10, 9, 11, 8, 8, 9, 9, 12, 10, 11, 8, 12, 9, 10, 11, 11, 11, 10, 12])
wt.freeze_panes = f"E{r1}"
wt.row_dimensions[TR].height = 40


# ======================= Cov (60D daily covariance, universe order) =======================
wcv = wb.create_sheet("Cov")
wcv.cell(row=1, column=1, value=f"60D daily return covariance (BBAS 3M window), {NC} universe names, as of {ASOF}. Row 2 = portfolio weights pulled from PreTrade_Check (SUMIFS). Static values; rebuild with build_risk.py. obs<40 names: diag = median var, off-diag 0.").font = f_bold
wcv.cell(row=3, column=1, value="Code").font = f_hdr; wcv.cell(row=3, column=2, value="Name").font = f_hdr
wcv.cell(row=2, column=2, value="Weight (from PreTrade) →").font = f_bold
nm = u.set_index("Code").Name
for j, c in enumerate(ccodes):
    col = 3 + j
    wcv.cell(row=3, column=col, value=c).font = f_bold
for i, c in enumerate(ccodes):
    row = 4 + i
    wcv.cell(row=row, column=1, value=c).font = f_norm; wcv.cell(row=row, column=2, value=nm.get(c, "")).font = f_norm
    vals = cov.values[i]
    for j in range(NC):
        wcv.cell(row=row, column=3 + j, value=float(vals[j]))
CFIRST, CLAST = L(3), L(3 + NC - 1)
CROWL, CROWF = 4, 4 + NC - 1
COV_MAT = f"Cov!${CFIRST}${CROWL}:${CLAST}${CROWF}"
COV_CODES = f"Cov!$A${CROWL}:$A${CROWF}"
COV_W = f"Cov!${CFIRST}$2:${CLAST}$2"
for j, c in enumerate(ccodes):
    cell = wcv.cell(row=2, column=3 + j, value=f'=SUMIFS(PreTrade_Check!$C${r1}:$C${rN},PreTrade_Check!$A${r1}:$A${rN},{L(3+j)}$3)')
    cell.number_format = PCT; cell.font = f_green
wcv.freeze_panes = "C4"
wcv.column_dimensions["A"].width = 9; wcv.column_dimensions["B"].width = 14

# ======================= DART_Verify =======================
wd = wb.create_sheet("DART_Verify", 3)
wd.cell(row=1, column=1, value=f"DART 검증 우선순위 — 이벤트 태그 {len(dart)}종목 (모델 지식 기반, 완료/무산 이벤트 혼재 가능). P1 = 아이디어 사용 또는 (actionable 이벤트 + 정량 시그널 확인) / P2 = actionable이나 정량 시그널 없음(종료 가능성) / P3 = 구조적·반복 테마(최근 공시만 확인). 노란 열은 검증 결과 입력.").font = f_bold
dh = ["Priority", "Code", "Name", "Tier", "Event Type", "Event Note (K)", "Corroborating signals (data)", "Idea #", "DART filing to check", "Flag", "Mcap (억)", "ADV20 (억)", "MaxPos Alpha@500억", "Verify status (input)", "Verified date (input)", "Outcome / next event (input)"]
for j, h in enumerate(dh, 1):
    wd.cell(row=3, column=j, value=h)
style_header(wd, 3, len(dh))
for i, (_, r) in enumerate(dart.iterrows(), 4):
    vals = [int(r.Priority), r.Code, r.Name, r.Tier, r.Event_Type, r.Event_Note, r.Signals, r.Idea, r.Filing, r.Flag, float(r.Mcap), float(r.ADV20), float(r.MaxPos), "", "", ""]
    for j, v in enumerate(vals, 1):
        c = wd.cell(row=i, column=j, value=v); c.font = f_norm; c.border = border; c.alignment = Alignment(wrap_text=True, vertical="top")
        if j == 11: c.number_format = NUM0
        if j == 12: c.number_format = NUM1
        if j == 13: c.number_format = PCT
        if j >= 14: c.fill = fill_yel; c.font = f_blue
wd.conditional_formatting.add(f"A4:A{3+len(dart)}", FormulaRule(formula=["A4=1"], fill=PatternFill("solid", fgColor="F8CBAD")))
wd.conditional_formatting.add(f"A4:A{3+len(dart)}", FormulaRule(formula=["A4=2"], fill=PatternFill("solid", fgColor="FFE699")))
setw(wd, [8, 8, 16, 6, 13, 44, 26, 7, 34, 34, 10, 9, 10, 14, 12, 26])
wd.freeze_panes = "D4"; wd.auto_filter.ref = f"A3:{L(len(dh))}{3+len(dart)}"
wd.row_dimensions[3].height = 36

# ======================= Alpha_Risk =======================
wa = wb.create_sheet("Alpha_Risk", 4)
wa.cell(row=1, column=1, value="Expected alpha vs risk budget — 아이디어별. 파란색 = 가정 입력(upside/downside/확률/기간). Risk = 60D 공분산 기반(Cov 시트): Solo VaR = 아이디어 단독 포트 VaR, Component VaR = 샘플 포트 내 기여분. Alpha/VaR = 연환산 NAV 기여 ÷ VaR.").font = f_bold
ah = ["#", "Idea", "Sleeve", "Notional (long-leg %NAV)", "Gross %NAV", "Upside (input)", "Downside (input)", "P(upside) (input)", "Horizon (m, input)", "E[ret] on notional", "NAV contribution", "NAV contribution p.a.", "Standalone VaR sum", "Solo VaR (cov)", "Component VaR in sample port", "Hedge benefit (standalone − solo)", "Alpha p.a. / Solo VaR", "Alpha p.a. / Component VaR", "Research ROI (NAV p.a. ≥ Params min?)"]
for j, h in enumerate(ah, 1):
    wa.cell(row=3, column=j, value=h)
style_header(wa, 3, len(ah))
for i, (_, r) in enumerate(irisk.iterrows(), 4):
    vals = [int(r.n), r.idea, r.sleeve, float(r.notional), float(r.gross), float(r.up), float(r.dn), float(r.p), int(r.hz)]
    for j, v in enumerate(vals, 1):
        c = wa.cell(row=i, column=j, value=v); c.font = f_norm; c.border = border
        if j in (4, 5): c.number_format = PCT
        if j in (6, 7, 8): c.number_format = PCT; c.font = f_blue; c.fill = fill_yel
        if j == 9: c.font = f_blue; c.fill = fill_yel
    fx = [(f"=H{i}*F{i}+(1-H{i})*G{i}", PCT), (f"=D{i}*J{i}", "0.00%"), (f"=K{i}*12/I{i}", "0.00%"),
          (float(r.standalone), "0.00%"), (float(r.solo_var), "0.00%"), (float(r.comp), "0.00%"), (f"=M{i}-N{i}", "0.00%"),
          (f'=IF(N{i}>0,L{i}/N{i},"n/a")', "0.00x"), (f'=IF(O{i}>0.0001,L{i}/O{i},"n/a (hedge)")', "0.00x"), (f'=IF(L{i}>={PMINNAV},"OK","Low")', "General")]
    for k, (f, fmt) in enumerate(fx, 10):
        c = wa.cell(row=i, column=k, value=f); c.font = f_norm; c.border = border; c.number_format = fmt
ar = 4 + len(irisk) + 2
wa.cell(row=ar, column=1, value="해석").font = f_bold
notes = ["E[ret]은 upside/downside 시나리오 가정의 확률가중 평균이며 모델링 값이 아님. 입력을 바꾸면 ROI 열이 재계산됨.",
         "Solo VaR = 아이디어 leg만으로 구성한 포트의 1D 99% VaR (60D 공분산). Pair의 hedge benefit = Standalone 합 − Solo VaR.",
         "Component VaR가 ≈0 또는 음수면 해당 아이디어가 샘플 포트에서 hedge 역할(리스크 소비 없음). obs<40 종목(신규상장)은 공분산 불안정 → Universe 시트 Cov flag 참조.",
         "Research ROI: 연환산 NAV 기여가 Params B26(0.5%p) 미만이면 Low. 1% 포지션 아이디어는 사실상 모두 Low → Event sleeve에서만, 리서치 시간 최소화."]
for k, t in enumerate(notes, 1):
    wa.cell(row=ar + k, column=1, value=t).font = f_norm
setw(wa, [4, 30, 7, 12, 9, 9, 9, 9, 9, 10, 10, 10, 10, 10, 12, 12, 11, 12, 14])
wa.freeze_panes = "C4"; wa.row_dimensions[3].height = 48

# ======================= Excluded =======================
we = wb.create_sheet("Excluded")
we.cell(row=1, column=1, value=f"Exclusion summary (전체 상장 {len(u)}종목, as of {ASOF})").font = f_bold
reasons = u["Excl_Reason"].replace("", "(포함 또는 유동성 부족)").value_counts()
for j, h in enumerate(["Reason", "# names"], 1):
    we.cell(row=3, column=j, value=h)
style_header(we, 3, 2)
r = 4
for k, v in reasons.items():
    we.cell(row=r, column=1, value=k).font = f_norm; c = we.cell(row=r, column=2, value=int(v)); c.font = f_norm; c.number_format = NUM0; r += 1
t2 = u["Tier"].value_counts()
r += 1
for j, h in enumerate(["Tier", "# names"], 1):
    we.cell(row=r, column=j, value=h)
style_header(we, r, 2); r += 1
for k, v in t2.items():
    we.cell(row=r, column=1, value=k).font = f_norm; c = we.cell(row=r, column=2, value=int(v)); c.font = f_norm; c.number_format = NUM0; r += 1
r += 2
we.cell(row=r, column=1, value="시총 < 3,000억이지만 ADV20 ≥ 30억 (유동성은 충분, BBAS ⑥로 편입 불가 — 3천억 미만 Event Play 허용 여부는 미팅 확인 사항 P7)").font = f_bold
r += 1
ca = u[(u.Tier.str.startswith("C")) & (u.ADV20 >= 30 * E) & (~u.Is_SPAC)].sort_values("ADV20", ascending=False).head(60)
cdf = pd.DataFrame({"Code": ca.Code, "Name": ca.Name, "Market": ca.Market, "Mcap (억)": ca.Mcap / E, "ADV20 (억)": ca.ADV20 / E, "Ret YTD": ca.Ret_YTD, "Ret 1M": ca.Ret_1M, "Vol spike": ca.VolSpike, "Event Type (K)": ca.Event_Type, "Quant Flags": ca.Quant_Flags})
r = write_df(we, cdf, r, fmts={"Mcap (억)": NUM0, "ADV20 (억)": NUM1, "Ret YTD": PCT, "Ret 1M": PCT, "Vol spike": MULT}) + 3
we.cell(row=r, column=1, value="X Illiquid: 시총 ≥ 3,000억이지만 Event sleeve(10일) 기준으로도 1% 사이징 불가 (ADV_min < 2.5억) → 유니버스 제외").font = f_bold
r += 1
cb = u[(u.Tier == "X Illiquid (<1% @10d)")].sort_values("Mcap", ascending=False)
cdf = pd.DataFrame({"Code": cb.Code, "Name": cb.Name, "Market": cb.Market, "Mcap (억)": cb.Mcap / E, "ADV20 (억)": cb.ADV20 / E, "MaxPos %NAV @500억": cb.MaxPosPct_500, "DTL 2% @500억 (d)": cb.DTL2_500, "Ret YTD": cb.Ret_YTD, "Event Type (K)": cb.Event_Type, "Event Note": cb.Event_Note})
r = write_df(we, cdf, r, fmts={"Mcap (억)": NUM0, "ADV20 (억)": NUM1, "MaxPos %NAV @500억": PCT, "DTL 2% @500억 (d)": NUM1, "Ret YTD": PCT}) + 3
we.cell(row=r, column=1, value="L4 Event-only: Alpha sleeve(5일) 기준 1% 미만, Event sleeve(10일) 기준 1% 이상 → 사전승인 Event Play로만 편입").font = f_bold
r += 1
cc = u[(u.Tier == "L4 Event-only (≥1% @10d)")].sort_values("Mcap", ascending=False)
cdf = pd.DataFrame({"Code": cc.Code, "Name": cc.Name, "Market": cc.Market, "Mcap (억)": cc.Mcap / E, "ADV20 (억)": cc.ADV20 / E, "MaxPos %NAV @500억": cc.MaxPosPct_500, "Ret YTD": cc.Ret_YTD, "Event Type (K)": cc.Event_Type})
write_df(we, cdf, r, fmts={"Mcap (억)": NUM0, "ADV20 (억)": NUM1, "MaxPos %NAV @500억": PCT, "Ret YTD": PCT})
setw(we, [40, 18, 10, 12, 12, 12, 12, 10, 16, 50])

# ======================= CIQ_Fields =======================
wc = wb.create_sheet("CIQ_Fields")
wc.cell(row=1, column=1, value="Capital IQ Excel Plug-in 수식 정의 (Universe 시트). 형식: =CIQ(ticker, mnemonic). Ticker: KOSE:A005930 (KOSPI) / KOSDAQ:A247540 (KOSDAQ). mnemonic은 CIQ Formula Builder에서 재확인.").font = f_bold
rows = [("P/E LTM", "IQ_PE_EXCL", "PE_RATIO", "LTM P/E excl. extraordinary"), ("P/E NTM", "IQ_PE_EXCL_FWD_CIQ", "BEST_PE_RATIO", "CIQ consensus NTM"),
        ("P/BV", "IQ_PBV", "PX_TO_BOOK_RATIO", "Latest"), ("TEV/EBITDA LTM", "IQ_TEV_EBITDA", "EV_TO_T12M_EBITDA", ""),
        ("ROE %", "IQ_RETURN_EQUITY", "RETURN_COM_EQY", "LTM"), ("ROIC %", "IQ_RETURN_CAPITAL", "RETURN_ON_CAP", "CIQ 'Return on Capital' ≈ ROIC; 세후 기준 확인"),
        ("Rev growth 1Y %", "IQ_TOTAL_REV_1YR_ANN_GROWTH", "SALES_GROWTH", ""), ("EPS growth 1Y %", "IQ_EPS_1YR_ANN_GROWTH", "EPS_GROWTH", ""),
        ("Div yield %", "IQ_DIVIDEND_YIELD", "EQY_DVD_YLD_IND", ""), ("Net debt", "IQ_NET_DEBT", "NET_DEBT", "KRW; 억 환산 필요"),
        ("Market cap", "IQ_MARKETCAP", "CUR_MKT_CAP", "검증용"), ("Avg daily value 3M", "IQ_AVG_DAILY_VALUE_3MO (확인)", "TURNOVER (or AVERAGE_VOLUME_30D*PX)", "ADV 교차검증"),
        ("Treasury shares %", "IQ_TREASURY_STOCK / IQ_TOTAL_SHARES (확인)", "", "자사주 소각 basket용"), ("Screen (권장)", "CIQ Screening: Country=Korea, Exchange=KOSE/KOSDAQ, Mkt cap ≥ KRW 300bn, 3M avg daily value ≥ KRW 1.67bn", "", "본 Universe 재현")]
for j, h in enumerate(["Field", "CIQ mnemonic", "Bloomberg equivalent (BDP)", "Note"], 1):
    wc.cell(row=3, column=j, value=h)
style_header(wc, 3, 4)
for i, rr in enumerate(rows, 4):
    for j, v in enumerate(rr, 1):
        c = wc.cell(row=i, column=j, value=v); c.font = f_norm; c.border = border
setw(wc, [22, 60, 36, 50])

out = f"{OUTDIR}/Korea_Universe_BBAS_{ASOF}.xlsx"
wb.save(out)
u.to_pickle(f"{OUTDIR}/universe.pkl")
print("saved", out, "universe", len(U), "special", (U.SpecialSit == "Y").sum())
