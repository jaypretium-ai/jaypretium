import pandas as pd, numpy as np, json
from openpyxl import load_workbook
OUT = "/tmp/claude-0/-home-user-jaypretium/06541f7f-96c0-5594-af5f-280151700bf4/scratchpad"
u = pd.read_pickle(f"{OUT}/universe.pkl"); p = pd.read_pickle(f"{OUT}/prefs.pkl"); E = 1e8
U = u[u.In_Universe].copy()
P = u.attrs["params"]
def f0(x): return "" if pd.isna(x) else f"{x:,.0f}"
def f1(x): return "" if pd.isna(x) else f"{x:,.1f}"
def pc(x): return "" if pd.isna(x) else f"{x*100:+.0f}%"
def pcu(x): return "" if pd.isna(x) else f"{x*100:.1f}%"
d = {}
d["asof"] = u.attrs["asof"]; d["n_all"] = len(u); d["n_univ"] = len(U)
d["n_eligible_mcap"] = int(((u.Tier != "X Excluded") & (u.Mcap_1M >= P["MCAP_MIN"])).sum())
d["n_illiquid"] = int((u.Tier == "X Illiquid (<1% @10d)").sum())
d["n_special"] = int((U.SpecialSit == "Y").sum()); d["n_event"] = int((U.Event_Type != "").sum())
tiers = ["L1 Full size (6% @5d)", "L2 Mid size (3-6% @5d)", "L3 Small size (1-3% @5d)", "L4 Event-only (≥1% @10d)"]
rows = []
for t in tiers + ["TOTAL"]:
    g = U if t == "TOTAL" else U[U.Tier == t]
    rows.append([t, f0(len(g)), f0(g.Mcap.median() / E), f0(g.ADV20.median() / E), pcu(g.Vol60_ann.median()), pc(g.Ret_YTD.median()),
                 f0((g.MaxPos_500_core >= 0.06 - 1e-9).sum()), f0((g.MaxPos_500_alpha >= 0.06 - 1e-9).sum()), f0((g.MaxPos_1000_alpha >= 0.06 - 1e-9).sum()), f0((g.SpecialSit == "Y").sum())])
d["tier_table"] = rows
d["tier_n"] = {t: int((U.Tier == t).sum()) for t in tiers}
for book in ["500", "1000"]:
    for sl in ["core", "alpha", "event"]:
        col = U[f"MaxPos_{book}_{sl}"]
        d[f"n6_{book}_{sl}"] = int((col >= 0.06 - 1e-9).sum()); d[f"n3_{book}_{sl}"] = int((col >= 0.03).sum()); d[f"n1_{book}_{sl}"] = int((col >= 0.01).sum())
        d[f"cap_{book}_{sl}"] = float(col.sum())
    d[f"n6_{book}_alpha_trough"] = int((U[f"MaxPos_{book}_alpha_Trough"] >= 0.06 - 1e-9).sum())
    d[f"n1_{book}_alpha_trough"] = int((U[f"MaxPos_{book}_alpha_Trough"] >= 0.01).sum())
d["mcap_share_L1"] = float(U[U.Tier == tiers[0]].Mcap.sum() / U.Mcap.sum())
d["univ_mcap_share"] = float(U.Mcap.sum() / u.Mcap.sum())
d["top10_share"] = float(U.sort_values("Mcap", ascending=False).head(10).Mcap.sum() / U.Mcap.sum())
d["capw_ytd"] = float((U.Ret_YTD * U.Mcap).sum() / U.Mcap[U.Ret_YTD.notna()].sum())
d["n_ytd100"] = int((U.Ret_YTD > 1).sum()); d["n_ytdm30"] = int((U.Ret_YTD < -0.3).sum())
d["n_vol15"] = int((U.Vol15_40D >= 2).sum()); d["n_vol15_L1"] = int(((U.Vol15_40D >= 2) & (U.Tier == tiers[0])).sum())
d["excl"] = {k: int(v) for k, v in u.Excl_Reason.replace("", "(시총 통과)").value_counts().items()}
d["bucket"] = {b: [int((u.Mcap_Bucket == b).sum()), int(((u.Mcap_Bucket == b) & u.In_Universe).sum())] for b in ["A ≥7000억", "B 3000-7000억", "C <3000억"]}
d["n_small_liquid"] = int(((u.Tier.str.startswith("C")) & (u.ADV20 >= 30 * E) & (~u.Is_SPAC)).sum())
d["n_L4"] = d["tier_n"][tiers[3]]
d["med_stab"] = float(U.ADV_Stability.median()); d["share_stab_lt50"] = float((U.ADV_Stability < 0.5).mean())
d["stab_by_tier"] = {t: float(U[U.Tier == t].ADV_Stability.median()) for t in tiers}
d["var_med"] = float(U.VaR99_1D.median()); d["var_med_L1"] = float(U[U.Tier == tiers[0]].VaR99_1D.median())
d["vol60_med_ann"] = float(U.Vol60_ann.median())
S = U[(U.SpecialSit == "Y") & (U.Event_Type != "")].sort_values("Mcap", ascending=False)
d["special_event"] = [[r.Name, r.Code, r.Tier.split(" ")[0], r.Event_Type, r.Event_Note, f0(r.Mcap / E), f0(r.ADV20 / E), pc(r.Ret_YTD), pcu(r.MaxPos_500_alpha)] for _, r in S.iterrows()]
Q = U[(U.SpecialSit == "Y") & (U.Event_Type == "")].sort_values("Mcap", ascending=False).head(25)
d["special_quant"] = [[r.Name, r.Code, r.Tier.split(" ")[0], r.Quant_Flags, f0(r.Mcap / E), f0(r.ADV20 / E), pc(r.Ret_YTD), pc(r.Ret_1M)] for _, r in Q.iterrows()]
pp = p[p.PrefADV20 >= 5 * E].sort_values("Discount").head(14)
tier_of = u.set_index("Code").Tier
d["prefs"] = [[r.Pref, r.Common, pc(r.Discount), f0(r.PrefADV20 / E), tier_of[r.PrefCode].split(" ")[0], tier_of[r.CommonCode].split(" ")[0]] for _, r in pp.iterrows()]
d["n_pref_pairs"] = int((p.PrefADV20 >= 5 * E).sum()); d["pref_med_disc"] = float(p[p.PrefADV20 >= 5 * E].Discount.median())
mc = u.set_index("Code")
hold = [("402340", "SK스퀘어", [("000660", 0.201)]), ("028260", "삼성물산", [("005930", 0.0501), ("207940", 0.4306), ("0126Z0", 0.4306), ("032830", 0.1934), ("018260", 0.1708)]), ("000150", "두산", [("034020", 0.304), ("454910", 0.682)]), ("003550", "LG", [("051910", 0.333), ("066570", 0.337), ("051900", 0.34), ("032640", 0.377), ("064400", 0.4995)]), ("034730", "SK", [("096770", 0.559), ("017670", 0.306), ("402340", 0.306), ("011790", 0.406), ("001740", 0.439), ("326030", 0.64)]), ("267250", "HD현대", [("009540", 0.37), ("267260", 0.37), ("443060", 0.62), ("267270", 0.30)]), ("180640", "한진칼", [("003490", 0.261)]), ("000670", "영풍", [("010130", 0.254)]), ("001040", "CJ", [("097950", 0.446), ("035760", 0.421)]), ("004990", "롯데지주", [("011170", 0.253), ("023530", 0.40), ("005300", 0.45), ("280360", 0.48)]), ("002790", "아모레퍼시픽홀딩스", [("090430", 0.372)]), ("001800", "오리온홀딩스", [("271560", 0.374)]), ("008930", "한미사이언스", [("128940", 0.414)]), ("005440", "현대지에프홀딩스", [("069960", 0.37), ("453340", 0.40)]), ("004800", "효성", [("298040", 0.325), ("298020", 0.203)])]
hr = []
for hc, hn, subs in hold:
    nav = sum(mc.loc[sc, "Mcap"] * st for sc, st in subs if sc in mc.index); m = mc.loc[hc, "Mcap"]
    hr.append([hn, mc.loc[hc, "Tier"].split(" ")[0], f0(m / E), f0(nav / E), f"{nav/m:.2f}x", pc(1 - m / nav), f0(mc.loc[hc, "ADV20"] / E)])
d["holdco"] = sorted(hr, key=lambda r: -float(r[4][:-1]))
for c in ["009155", "009150", "005387", "005380", "402340", "000660", "028260", "0126Z0", "207940", "0220W0", "000880", "089860", "267270", "002990", "047040", "006340", "000500", "005935", "005930", "010130", "008930", "128940", "035420", "026960", "012630"]:
    r = mc.loc[c]; d["nm_" + c] = dict(name=r.Name, tier=r.Tier.split(" ")[0], mcap=f0(r.Mcap / E), adv=f0(r.ADV20 / E), ytd=pc(r.Ret_YTD), m1=pc(r.Ret_1M), maxpos=pcu(r.MaxPos_500_alpha), maxpos_ev=pcu(r.MaxPos_500_event), dtl6=f1(r.DTL6_500), vol15=int(r.Vol15_40D), spike=f"{r.VolSpike:.1f}x", close=f0(r.Close), h52=pc(r.Pct_from_52wH))
d["disc_009155"] = pc(mc.loc["009155", "Close"] / mc.loc["009150", "Close"] - 1); d["disc_005387"] = pc(mc.loc["005387", "Close"] / mc.loc["005380", "Close"] - 1); d["disc_005935"] = pc(mc.loc["005935", "Close"] / mc.loc["005930", "Close"] - 1)
d["sq_cov"] = f"{mc.loc['000660','Mcap']*0.201/mc.loc['402340','Mcap']:.2f}x"
wb = load_workbook(f"{OUT}/Korea_Universe_BBAS_{d['asof']}.xlsx", data_only=True); ws = wb["PreTrade_Check"]
vals = {}
for r in ws.iter_rows(min_row=46, max_row=70, values_only=True):
    if r[0] and r[1] is not None: vals[r[0]] = r[1]
d["pt"] = {k: (float(v) if isinstance(v, (int, float)) else v) for k, v in vals.items()}
json.dump(d, open(f"{OUT}/memo_data.json", "w"), ensure_ascii=False, indent=1)
print({k: v for k, v in d.items() if isinstance(v, (int, float, str)) and not k.startswith("nm_")})
print(d["pt"])
